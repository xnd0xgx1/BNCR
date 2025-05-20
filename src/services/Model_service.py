from src.interfaces.di_interface import DocIntInterface
from src.interfaces.aoi_interface import AOIInterface
from src.interfaces.st_interface import STInterface
from src.interfaces.cosmos_interface import CosmosInterface
import logging
import json
from openpyxl import load_workbook
from datetime import datetime
from io import BytesIO
class ModelService:

    def __init__(self,azure_di:DocIntInterface,azure_oi: AOIInterface,azure_st: STInterface,azure_cosmos: CosmosInterface,base_url:str):
        self.azure_di = azure_di
        self.azure_oi = azure_oi
        self.azure_st = azure_st
        self.base_url = base_url
        self.azure_cosmos = azure_cosmos

    def process(self,pdf_files):
        contents = []
        confidences = []

        for file_name, file in pdf_files.items():
            logging.info(f"Processing file: {file.filename}")
            file_bytes = file.stream.read()
            file_stream = BytesIO(file_bytes)
            diresult = self.azure_di.Process(filestream=file_stream)

            contents.append(diresult["content"])
            confidences.append(diresult["average_page_confidence"])
            logging.error(f"Confidences acumuladas: {confidences}")
        
        # # for pdf in pdfbytes:
        # #     diresult = self.azure_di.Process(filestream=pdf)
        # #     logging.warning("Apending results to Content")
        # #     contents.append(diresult["content"])
        # #     confidence.append(diresult["average_page_confidence"])
        # #     logging.error(f"Contenido de conficencia: {confidence}")

        resultadotextocompleto = ""
        for content in contents:
            resultadotextocompleto += content  
          

        logging.info(f"Texto completo resultante (longitud {len(resultadotextocompleto)}):")
        result = self.azure_oi.Call(content=resultadotextocompleto)
        # logging.warning(f"INFORECIVED: {result}")
        logging.warning("Procesando valores para excel")
        periodos = sorted(result["periodos"], key=lambda p: p["año"])[:4]  # antiguos a recientes, máximo 4

        wb = load_workbook("Estructura para Estados Financieros IA (2).xlsx")
        hoja_res = wb["Resultados"]
        hoja_bal = wb["Balance"]

        def escribir_valores(hoja, data_dict, dest_col):
            # Prepara lista de rangos merged para la hoja
            merged_bounds = [
                (rng.min_row, rng.max_row, rng.min_col, rng.max_col)
                for rng in hoja.merged_cells.ranges
            ]
            def es_merged(row, col):
                return any(
                    mn_row <= row <= mx_row and mn_col <= col <= mx_col
                    for (mn_row, mx_row, mn_col, mx_col) in merged_bounds
                )

            for fila in range(1, hoja.max_row + 1):
                nombre = hoja.cell(row=fila, column=2).value  # columna B
                if nombre in data_dict and not es_merged(fila, dest_col):
                    hoja.cell(row=fila, column=dest_col).value = data_dict[nombre]

        # Itera sobre los periodos y columnas C(3), D(4), E(5), F(6)
        for idx, periodo in enumerate(periodos):
            col = 3 + idx
            # Opcional: poner el año en la fila 1 de esa columna
            hoja_res.cell(row=1, column=col).value = periodo["año"]
            hoja_bal.cell(row=1, column=col).value = periodo["año"]
            # Escribe los valores
            escribir_valores(hoja_res, periodo["estadoResultados"], dest_col=col)
            for sec in periodo["balanceGeneral"].values():
                escribir_valores(hoja_bal, sec, dest_col=col)
        logging.warning("Almacenando datos")
        logging.warning("Guardando Excel")
      
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"{fecha_actual}.xlsx"

        # Guardar el archivo en el storage
        self.azure_st.Save(nombre_archivo,wb)

        # Agregar resultados
        result["Calidad"] = confidences
        result["URL"] = f"{self.base_url}/api/excel?nombre={nombre_archivo}"    

        #Guardar resultados en cosmos
        
      
        item_creado = self.azure_cosmos.save_record(result)
        print("Item creado:", item_creado)


        # wb.save("resultadoExcel2.xlsx")
        return json.dumps(result)
    def getdocument(self,name):
        result = self.azure_st.Get(document_name=name)
        return result
